#!/usr/bin/python3

import getpass
import openpyxl
import os
import smtplib, ssl
import xlsxwriter

from buildingFloorEntry import buildingFloorEntry
from buildingFloorEntryUtil import createBuildingFloorEntryObject
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


def generateOutputFile(building_floor_dict, material_dict):
    rows_to_write = []
    for key in building_floor_dict:
        team_name = key.split('_')[2]
        building_floor_name = "Building " + key.split('_')[0] + " Floor " + key.split('_')[1]
        total_cost = 0.0
        for entry in building_floor_dict[key]:
            cost = material_dict[entry[0]]
            total_cost += cost
        utilization_percentage = total_cost / entry[1] * 100.0
        utilization_percentage_string = "{:.2f}".format(utilization_percentage) + "%"
        rows_to_write.append([team_name, building_floor_name, utilization_percentage_string])
    try:
        os.remove('data/output.xlsx')
    except OSError:
        pass
    workbook = xlsxwriter.Workbook('data/output.xlsx')
    worksheet = workbook.add_worksheet("Team plan sheet")
    title = ['team', 'building and floor', 'budget utilization']
    for i in range(0,3):
        worksheet.write(0, i, title[i])
    row_index = 1
    for entry in rows_to_write:
        for col in range(len(entry)):
            worksheet.write(row_index, col, entry[col])
        row_index += 1
    workbook.close()

def sendEmail():
    port = 465  # For SSL
    sender_email = "davidlucs@gmail.com"
    receiver_email = "davidlucs@gmail.com"
    message = MIMEMultipart("alternative")
    message["Subject"] = "Building task report"
    message["From"] = sender_email
    message["To"] = receiver_email


    html = """\
    <html>
    <body>
        <p>Hi,<br>
        Attached is the building plan document , let me know if you have any question! <br>
        <br>
        Sincerely, <br>
        Yanting
        </p>
    </body>
    </html>
    """
    html_text = MIMEText(html, "html")
    message.attach(html_text)
    filename = "data/output.xlsx"  # Use a hardcoded name here for testing purpose
    with open(filename, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )
    message.attach(part)
    text = message.as_string()

    password = getpass.getpass(prompt = "Type your gmail app password and press enter: ")
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
        server.login("davidlucs@gmail.com", password)
        server.sendmail(sender_email, receiver_email, message.as_string())
        server.close()
    print ('successfully sent the mail')

def readBuildingFloorFile():
    # Load the building floor entry file
    building_workbook = openpyxl.load_workbook("data/building.xlsx")
    building_floor_entry_list = []
    # Get active sheet
    building_worksheet = building_workbook.active

    # Loop through each row and column to read all entries
    for row in range(0, building_worksheet.max_row):
        temp_entry = buildingFloorEntry('none',0,'none',0,0,0,'N/A')
        col_counter = 0
        for col in building_worksheet.iter_cols(0, building_worksheet.max_column):
            temp_entry = createBuildingFloorEntryObject(col_counter, col[row].value, temp_entry)
            col_counter += 1
            print(col[row].value, end="\t\t")
        building_floor_entry_list.append(temp_entry)
        print('')

    building_floor_dict = {}
    for i in range(1, len(building_floor_entry_list)):
        key = f"{building_floor_entry_list[i].name}_{building_floor_entry_list[i].floor}_{building_floor_entry_list[i].team}"
        temp_tuple = (building_floor_entry_list[i].material, building_floor_entry_list[i].budget, building_floor_entry_list[i].getNeedAmount())
        if building_floor_dict.__contains__(key):
            building_floor_dict[key].append(temp_tuple)
        else:
            building_floor_dict[key] = [temp_tuple]

    return building_floor_dict

def readMaterialFile():
    material_workbook = openpyxl.load_workbook("data/material.xlsx")
    material_cost_map = {}
    material_worksheet = material_workbook.active

    # Loop through each row and column to read all entries
    for row in range(0, material_worksheet.max_row):
        material_list = []
        for col in material_worksheet.iter_cols(0, material_worksheet.max_column):
            print(col[row].value, end="\t\t")
            material_list.append(col[row].value)
        print('')
        material_cost_map[material_list[0]] = material_list[1]
    return material_cost_map
